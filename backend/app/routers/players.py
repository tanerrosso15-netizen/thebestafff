"""Oyuncular — filtreli liste (admin tümü, affiliate kendi oyuncuları) + Excel export."""
from __future__ import annotations

import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import get_current_user
from app.models import Affiliate, Player, User
from app.schemas import PlayerOut

router = APIRouter(prefix="/api/players", tags=["players"])


def _to_out(p: Player) -> PlayerOut:
    out = PlayerOut.model_validate(p)
    out.affiliate_name = p.affiliate.name if p.affiliate else None
    return out


def _build_query(
    db: Session,
    user: User,
    *,
    player_id: str | None,
    name: str | None,
    username: str | None,
    btag: str | None,
    min_balance: float | None,
    affiliate_id: int | None,
    date_from: str | None,
    date_to: str | None,
):
    query = db.query(Player)
    if user.role == "affiliate":
        own = db.query(Affiliate).filter(Affiliate.user_id == user.id).first()
        if not own:
            return None
        child_ids = [c.id for c in db.query(Affiliate).filter(Affiliate.parent_id == own.id).all()]
        query = query.filter(Player.affiliate_id.in_([own.id, *child_ids]))

    if affiliate_id is not None:
        query = query.filter(Player.affiliate_id == affiliate_id)
    if player_id:
        query = query.filter(Player.external_id.ilike(f"%{player_id}%"))
    if name:
        query = query.filter(Player.name.ilike(f"%{name}%"))
    if username:
        query = query.filter(Player.username.ilike(f"%{username}%"))
    if btag:
        query = query.filter(Player.btag.ilike(f"%{btag}%"))
    if min_balance is not None:
        query = query.filter(Player.balance >= min_balance)
    if date_from:
        try:
            query = query.filter(Player.registered_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Player.registered_at <= datetime.fromisoformat(date_to))
        except ValueError:
            pass
    return query


@router.get("")
def list_players(
    player_id: str | None = Query(None),
    name: str | None = Query(None),
    username: str | None = Query(None),
    btag: str | None = Query(None),
    min_balance: float | None = Query(None),
    affiliate_id: int | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=200),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = _build_query(
        db, user, player_id=player_id, name=name, username=username, btag=btag,
        min_balance=min_balance, affiliate_id=affiliate_id, date_from=date_from, date_to=date_to,
    )
    if query is None:
        return {"items": [], "total": 0, "page": page, "per_page": per_page}

    total = query.count()
    rows = (
        query.order_by(Player.registered_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    return {
        "items": [_to_out(p).model_dump() for p in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
    }


@router.get("/export")
def export_players(
    player_id: str | None = Query(None),
    name: str | None = Query(None),
    username: str | None = Query(None),
    btag: str | None = Query(None),
    min_balance: float | None = Query(None),
    affiliate_id: int | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Uygulanan filtrelere göre oyuncuları Excel (.xlsx) olarak indir."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    query = _build_query(
        db, user, player_id=player_id, name=name, username=username, btag=btag,
        min_balance=min_balance, affiliate_id=affiliate_id, date_from=date_from, date_to=date_to,
    )
    rows = [] if query is None else query.order_by(Player.registered_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Oyuncular"
    headers = [
        "ID", "Kullanıcı Adı", "Kategori", "btag", "Affiliate", "Para Birimi",
        "Bakiye", "Toplam Yatırım", "Toplam Çekim", "Kayıt Tarihi",
    ]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="7C3AED")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill

    for p in rows:
        ws.append([
            p.external_id or p.id,
            p.username or p.name or "",
            p.category or "",
            p.btag or "",
            p.affiliate.name if p.affiliate else "",
            p.currency or "",
            round(p.balance or 0, 2),
            round(p.deposit_total or 0, 2),
            round(p.withdrawal_total or 0, 2),
            p.registered_at.strftime("%Y-%m-%d %H:%M") if p.registered_at else "",
        ])

    widths = [14, 22, 16, 14, 20, 10, 14, 16, 16, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"pqp_oyuncular_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{player_id}", response_model=PlayerOut)
def get_player(player_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    p = db.query(Player).filter(Player.id == player_id).first()
    if not p:
        from fastapi import HTTPException

        raise HTTPException(status_code=404, detail="Oyuncu bulunamadı.")
    return _to_out(p)
